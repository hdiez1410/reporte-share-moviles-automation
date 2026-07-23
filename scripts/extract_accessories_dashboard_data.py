from __future__ import annotations

import argparse
import json
import math
import struct
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from extract_sales_report_data import (
    MONTH_ABBR,
    decode_rk,
    group_channel,
    iter_records,
    load_crosses,
    norm_key,
    normalized,
    parse_shared_strings,
    transaction_date,
)
from extract_sales_report_data_multi import identify_xlsb_sheets, month_label, week_from_date_unique


SALES_FIELDS = [
    "PERIODO",
    "DIA",
    "ESTADO",
    "GRUPO",
    "FAMILIA",
    "TIPOTRANSACCION",
    "MARCA",
    "CANAL",
    "PUNTODEVENTA",
    "MARCAMODELO",
    "TIPO",
    "SUBTIPO",
    "ITEMS",
]
STOCK_FIELDS = [
    "PERIODO",
    "DIA",
    "ESTADO",
    "USO",
    "FAMILIA",
    "MARCA",
    "CANAL",
    "PUNTODEVENTA",
    "MARCAMODELO",
    "TIPO",
    "SUBTIPO",
    "ITEMS",
]
STOCK_USO = {"Normal", "Pack"}


def numeric_value(value) -> float | None:
    if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
        return float(value)
    text = normalized(value)
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def category(value, fallback: str) -> str:
    text = normalized(value)
    return text.upper() if text else fallback


def header_indexes(header: list[str], fields: list[str]) -> dict[str, int]:
    index = {normalized(value): position for position, value in enumerate(header) if normalized(value)}
    missing = [field for field in fields if field not in index]
    if missing:
        raise RuntimeError(f"Columnas no encontradas: {', '.join(missing)}")
    return {field: index[field] for field in fields}


def iter_xlsx_records(path: Path, sheet: str, fields: list[str]):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return
    rows = wb[sheet].iter_rows(values_only=True)
    indexes = header_indexes([normalized(value) for value in next(rows)], fields)
    try:
        for row in rows:
            yield {field: row[position] if position < len(row) else None for field, position in indexes.items()}
    finally:
        wb.close()


def iter_binary_records(data: bytes, shared: list[str], fields: list[str]):
    header = []
    indexes = {}
    current_row = None
    row_values: dict[int, object] = {}

    def materialize(row_index, values):
        nonlocal header, indexes
        if row_index is None:
            return None
        if row_index == 0:
            header = [normalized(values.get(i)) for i in range(max(values.keys(), default=-1) + 1)]
            indexes = header_indexes(header, fields)
            return None
        return {field: values.get(position) for field, position in indexes.items()}

    for record_type, payload in iter_records(data):
        if record_type == 0:
            record = materialize(current_row, row_values)
            if record is not None:
                yield record
            current_row = int.from_bytes(payload[0:4], "little", signed=False)
            row_values = {}
        elif current_row is not None and record_type == 7 and len(payload) >= 12:
            column = int.from_bytes(payload[0:4], "little", signed=False)
            string_index = int.from_bytes(payload[8:12], "little", signed=False)
            row_values[column] = shared[string_index] if 0 <= string_index < len(shared) else ""
        elif current_row is not None and record_type == 5 and len(payload) >= 16:
            column = int.from_bytes(payload[0:4], "little", signed=False)
            row_values[column] = struct.unpack("<d", payload[8:16])[0]
        elif current_row is not None and record_type == 2 and len(payload) >= 12:
            column = int.from_bytes(payload[0:4], "little", signed=False)
            row_values[column] = decode_rk(payload[8:12])
    record = materialize(current_row, row_values)
    if record is not None:
        yield record


def iter_xlsb_records(path: Path, stock: bool):
    with zipfile.ZipFile(path) as archive:
        shared = parse_shared_strings(archive)
        sales_data, stock_data = identify_xlsb_sheets(archive, shared)
        data = stock_data if stock else sales_data
    if data is None:
        return
    yield from iter_binary_records(data, shared, STOCK_FIELDS if stock else SALES_FIELDS)


def source_records(path: Path, stock: bool):
    if path.suffix.lower() == ".xlsb":
        yield from iter_xlsb_records(path, stock)
    else:
        yield from iter_xlsx_records(path, "Base Stocks" if stock else "Base Movs", STOCK_FIELDS if stock else SALES_FIELDS)


def parse_source(value: str) -> tuple[str, Path]:
    period, separator, path = value.partition("=")
    if not separator or len(period) != 6 or not period.isdigit():
        raise argparse.ArgumentTypeError("Use PERIODO=/ruta/al/archivo, por ejemplo 202606=/ruta/base.xlsb")
    return period, Path(path)


def row_dimensions(record: dict, crosses: dict) -> tuple[str, str, str, str, str, str, str]:
    channel = group_channel(normalized(record.get("CANAL")) or "Sin canal")
    pdv = normalized(record.get("PUNTODEVENTA")) or "Sin punto de venta"
    regional = crosses["regional_map"].get(norm_key(pdv), "Sin regional")
    type_name = category(record.get("TIPO"), "SIN TIPO")
    subtype = category(record.get("SUBTIPO"), "SIN SUBTIPO")
    if type_name == "TABLET":
        subtype = "TABLET"
    brand = category(record.get("MARCA"), "SIN MARCA")
    model = normalized(record.get("MARCAMODELO")) or "Sin modelo"
    return channel, regional, pdv, type_name, subtype, brand, model


def build(sales_sources: list[tuple[str, Path]], stock_sources: list[tuple[str, Path]], cruces_path: Path) -> dict:
    crosses = load_crosses(cruces_path)
    sales = defaultdict(float)
    stocks = defaultdict(float)
    summaries = []
    missing_regional_sales = Counter()
    dates = set()
    stock_dates = set()
    filtered_rows = 0
    stock_filtered_rows = 0

    for target_period, path in sales_sources:
        print(json.dumps({"processing": "sales", "period": target_period, "file": path.name}, ensure_ascii=False), flush=True)
        source_rows = 0
        source_filtered = 0
        for record in source_records(path, stock=False):
            source_rows += 1
            if normalized(record.get("PERIODO")) != target_period:
                continue
            if normalized(record.get("ESTADO")) != "Nuevos":
                continue
            if normalized(record.get("GRUPO")) != "Consumo":
                continue
            if normalized(record.get("FAMILIA")) != "Accesorios":
                continue
            if normalized(record.get("TIPOTRANSACCION")) != "3.SALIDAS":
                continue
            items = numeric_value(record.get("ITEMS"))
            if items is None or abs(items) < 1e-12:
                continue
            date_value = transaction_date(record.get("PERIODO"), record.get("DIA"))
            week, _, _, _ = week_from_date_unique(date_value)
            dimensions = row_dimensions(record, crosses)
            date_text = date_value.isoformat() if date_value else "Sin fecha"
            sales[(*dimensions, week, date_text, target_period, target_period[:4])] += -items
            source_filtered += 1
            filtered_rows += 1
            if date_value:
                dates.add(date_value)
            if dimensions[1] == "Sin regional":
                missing_regional_sales[dimensions[2]] += -items
        summaries.append(
            {
                "file": str(path),
                "period": target_period,
                "kind": "sales",
                "source_rows": source_rows,
                "filtered_rows": source_filtered,
            }
        )
        print(json.dumps(summaries[-1], ensure_ascii=False), flush=True)

    stock_summaries = []
    for target_period, path in stock_sources:
        print(json.dumps({"processing": "stock", "period": target_period, "file": path.name}, ensure_ascii=False), flush=True)
        source_rows = 0
        source_filtered = 0
        for record in source_records(path, stock=True):
            source_rows += 1
            if normalized(record.get("PERIODO")) != target_period:
                continue
            if normalized(record.get("ESTADO")) != "Nuevos":
                continue
            if normalized(record.get("USO")) not in STOCK_USO:
                continue
            if normalized(record.get("FAMILIA")) != "Accesorios":
                continue
            items = numeric_value(record.get("ITEMS"))
            if items is None or abs(items) < 1e-12:
                continue
            date_value = transaction_date(record.get("PERIODO"), record.get("DIA"))
            dimensions = row_dimensions(record, crosses)
            date_text = date_value.isoformat() if date_value else "Sin fecha"
            stocks[(*dimensions, date_text, target_period, target_period[:4])] += items
            source_filtered += 1
            stock_filtered_rows += 1
            if date_value:
                stock_dates.add(date_value)
        summary = {
            "file": str(path),
            "period": target_period,
            "kind": "stock",
            "source_rows": source_rows,
            "filtered_rows": source_filtered,
        }
        summaries.append(summary)
        stock_summaries.append(summary)
        print(json.dumps(summary, ensure_ascii=False), flush=True)

    dashboard_data = []
    for key, units in sales.items():
        channel, regional, pdv, type_name, subtype, brand, model, week, date_text, period, year = key
        dashboard_data.append(
            {
                "CANAL_AGRUPADO": channel,
                "REGIONAL_HONOR": regional,
                "PUNTODEVENTA": pdv,
                "TIPO": type_name,
                "SUBTIPO": subtype,
                "SEGMENTO_HONOR": subtype,
                "MARCA": brand,
                "MARCAMODELO": model,
                "SEMANA": week,
                "FECHA": date_text,
                "DIA": date_text[-2:] if date_text != "Sin fecha" else "",
                "PERIODO": period,
                "ANIO": year,
                "MES_LABEL": month_label(period),
                "VENTAS": round(units, 6),
                "PRECIO": None,
            }
        )
    dashboard_data.sort(
        key=lambda row: (
            row["PERIODO"],
            row["CANAL_AGRUPADO"],
            row["REGIONAL_HONOR"],
            row["PUNTODEVENTA"],
            row["TIPO"],
            row["SUBTIPO"],
            row["MARCA"],
            row["MARCAMODELO"],
            row["FECHA"],
        )
    )

    stock_data = []
    for key, units in stocks.items():
        channel, regional, pdv, type_name, subtype, brand, model, date_text, period, year = key
        stock_data.append(
            {
                "CANAL_AGRUPADO": channel,
                "REGIONAL_HONOR": regional,
                "PUNTODEVENTA": pdv,
                "TIPO": type_name,
                "SUBTIPO": subtype,
                "SEGMENTO_HONOR": subtype,
                "MARCA": brand,
                "MARCAMODELO": model,
                "FECHA": date_text,
                "DIA": date_text[-2:] if date_text != "Sin fecha" else "",
                "PERIODO": period,
                "ANIO": year,
                "MES_LABEL": month_label(period),
                "STOCK": round(units, 6),
            }
        )
    stock_data.sort(key=lambda row: (row["FECHA"], row["CANAL_AGRUPADO"], row["PUNTODEVENTA"], row["TIPO"], row["MARCAMODELO"]))

    period_totals = defaultdict(float)
    brand_totals = defaultdict(float)
    weeks = {}
    for row in dashboard_data:
        period_totals[row["PERIODO"]] += row["VENTAS"]
        brand_totals[row["MARCA"]] += row["VENTAS"]
        if row["FECHA"] != "Sin fecha" and row["SEMANA"] not in weeks:
            date_value = datetime.strptime(row["FECHA"], "%Y-%m-%d").date()
            _, number, start, end = week_from_date_unique(date_value)
            weeks[row["SEMANA"]] = (number, start, end)

    latest_stock_date = max((row["FECHA"] for row in stock_data if row["FECHA"] != "Sin fecha"), default="Sin fecha")
    latest_stock = [row for row in stock_data if row["FECHA"] == latest_stock_date]
    periods = sorted(period_totals)
    years = sorted({row["ANIO"] for row in dashboard_data})
    brands = sorted(brand_totals, key=lambda value: (-brand_totals[value], value))
    types = sorted({row["TIPO"] for row in dashboard_data})
    subtypes = sorted({row["SUBTIPO"] for row in dashboard_data})
    channels = sorted(
        {row["CANAL_AGRUPADO"] for row in dashboard_data},
        key=lambda value: -sum(row["VENTAS"] for row in dashboard_data if row["CANAL_AGRUPADO"] == value),
    )

    return {
        "report_kind": "accessories",
        "source_file": "Bases Oracle 2026 - Familia Accesorios",
        "source_files": [str(path) for _, path in sales_sources],
        "source_file_summaries": summaries,
        "stock_source_file": str(stock_sources[-1][1]) if stock_sources else "",
        "stock_file_summary": stock_summaries[-1] if stock_summaries else {},
        "cruces_file": str(cruces_path),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "filters": {
            "ESTADO": "Nuevos",
            "GRUPO": "Consumo",
            "FAMILIA": "Accesorios",
            "TIPOTRANSACCION": "3.SALIDAS",
        },
        "metric_note": "VENTAS = -SUM(ITEMS). Stock = SUM(ITEMS) con ESTADO=Nuevos y USO Normal/Pack.",
        "week_note": "Semana calculada de lunes a domingo desde PERIODO + DIA.",
        "stock_note": f"Los indicadores de stock actual usan la foto más reciente ({latest_stock_date}).",
        "filtered_rows": filtered_rows,
        "stock_filtered_rows": len(latest_stock),
        "stock_total": round(sum(row["STOCK"] for row in latest_stock), 6),
        "total_sales": round(sum(row["VENTAS"] for row in dashboard_data), 6),
        "dates": [
            {
                "FECHA": value.isoformat(),
                "DIA": f"{value.day:02d}",
                "PERIODO": f"{value.year}{value.month:02d}",
                "ANIO": str(value.year),
                "LABEL": f"{value.day:02d}-{MONTH_ABBR[value.month]}",
            }
            for value in sorted(dates)
        ],
        "stock_dates": [
            {
                "FECHA": value.isoformat(),
                "DIA": f"{value.day:02d}",
                "PERIODO": f"{value.year}{value.month:02d}",
                "ANIO": str(value.year),
                "LABEL": f"{value.day:02d}-{MONTH_ABBR[value.month]}",
            }
            for value in sorted(stock_dates)
        ],
        "week_ranges": [
            {
                "SEMANA": week,
                "NUMERO": number,
                "ANIO": str(start.isocalendar().year),
                "DESDE": start.isoformat(),
                "HASTA": end.isoformat(),
                "LABEL": f"{start.isocalendar().year} S{number} ({start.day:02d}-{MONTH_ABBR[start.month]} al {end.day:02d}-{MONTH_ABBR[end.month]})",
                "SHORT": f"S{number}",
            }
            for week, (number, start, end) in sorted(weeks.items(), key=lambda item: item[1][1])
        ],
        "periods": [
            {"PERIODO": period, "ANIO": period[:4], "LABEL": month_label(period), "VENTAS": round(period_totals[period], 6)}
            for period in periods
        ],
        "years": years,
        "dashboard_data": dashboard_data,
        "stock_data": stock_data,
        "lists": {
            "years": years,
            "periods": periods,
            "period_labels": {period: month_label(period) for period in periods},
            "channels": channels,
            "regionals": sorted({row["REGIONAL_HONOR"] for row in dashboard_data}),
            "pdvs": sorted({row["PUNTODEVENTA"] for row in dashboard_data}),
            "segmentos_honor": subtypes,
            "types": types,
            "subtypes": subtypes,
            "models": sorted({row["MARCAMODELO"] for row in dashboard_data}),
            "brands": brands,
        },
        "cross_summary": {
            "missing_sales": {"REGIONAL": round(sum(missing_regional_sales.values()), 6)} if missing_regional_sales else {},
            "missing_regional_pdvs": [
                {"PUNTODEVENTA": pdv, "VENTAS": round(units, 6)}
                for pdv, units in missing_regional_sales.most_common(50)
            ],
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Construye la data del reporte web de Accesorios.")
    parser.add_argument("output")
    parser.add_argument("--cruces", required=True)
    parser.add_argument("--sales", action="append", type=parse_source, required=True)
    parser.add_argument("--stock", action="append", type=parse_source, required=True)
    args = parser.parse_args()
    data = build(args.sales, args.stock, Path(args.cruces))
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(output),
                "total_sales": data["total_sales"],
                "stock_total": data["stock_total"],
                "periods": data["lists"]["periods"],
                "types": data["lists"]["types"],
                "subtypes": len(data["lists"]["subtypes"]),
                "missing_regional_sales": data["cross_summary"].get("missing_sales", {}),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
