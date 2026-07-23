from __future__ import annotations

import argparse
import csv
import json
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from enrich_accessories_report_data import best_price_match, extract_price_candidates


CHANNEL_DISPLAY_ALIASES = {
    "Brightstar": "Remotos",
    "Brighstar": "Remotos",
    "Televenta": "Remotos",
    "TEX": "Tiendas Express",
    "TIENDAS EXPRESS": "Tiendas Express",
    "TEX SATELITE": "Tiendas Express",
    "Tiendas Propias Franquicias": "Tiendas",
}
SUPERVISOR_MAPPING_PATHS = [
    Path(__file__).resolve().parent / "supervisor_honor_mapping.json",
    Path(__file__).resolve().parents[1] / "outputs" / "cruces" / "supervisor_honor_mapping.json",
]
SUPERVISOR_CSV_PATH = Path(__file__).resolve().parent / "pdv_supervisor_honor.csv"


def dashboard_channel(value: str) -> str:
    return CHANNEL_DISPLAY_ALIASES.get(value, value)


def normalized_lookup_key(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.strip().upper().split())


def load_supervisor_mapping() -> dict[str, str]:
    supervisor_mapping: dict[str, str] = {}
    for mapping_path in SUPERVISOR_MAPPING_PATHS:
        if not mapping_path.exists():
            continue
        try:
            payload = json.loads(mapping_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            payload = {}
        mapping = payload.get("mapping", payload)
        if isinstance(mapping, dict):
            for key, value in mapping.items():
                supervisor = str(value or "Otros").strip() or "Otros"
                supervisor_mapping[normalized_lookup_key(key)] = "Otros" if supervisor.casefold() == "otros" else supervisor
    if SUPERVISOR_CSV_PATH.exists():
        with SUPERVISOR_CSV_PATH.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle, delimiter=";"):
                pdv = str(row.get("punto_de_venta") or "").strip()
                supervisor = str(row.get("Super") or "").strip()
                if not pdv:
                    continue
                if not supervisor or normalized_lookup_key(supervisor) == normalized_lookup_key(pdv):
                    supervisor = "Otros"
                elif supervisor.casefold() == "otros":
                    supervisor = "Otros"
                supervisor_mapping[normalized_lookup_key(pdv)] = supervisor
    return supervisor_mapping


def supervisor_for_pdv(pdv, supervisor_by_pdv: dict[str, str]) -> str:
    return supervisor_by_pdv.get(normalized_lookup_key(pdv), "Otros")


def period_label_map(periods: list[dict]) -> dict[str, str]:
    return {item["PERIODO"]: item.get("LABEL", item["PERIODO"]) for item in periods if item.get("PERIODO")}


def weighted_model_prices(rows: list[dict]) -> dict[str, float]:
    amount = defaultdict(float)
    units = defaultdict(float)
    for row in rows:
        price = row.get("PRECIO")
        if price is None:
            continue
        model = row["MARCAMODELO"]
        sales = float(row["VENTAS"])
        amount[model] += float(price) * sales
        units[model] += sales
    return {model: amount[model] / units[model] for model in sorted(amount) if units[model]}


def parse_price_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit() and len(text) >= 8:
        return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    return text[:10]


def latest_year(data: dict) -> str:
    years = [str(year) for year in data.get("years", []) if year]
    return sorted(years)[-1] if years else ""


def price_history_from_cruces(data: dict) -> dict:
    year = latest_year(data)
    cruces_file = Path(data.get("cruces_file", ""))
    empty = {"year": year, "source": str(cruces_file), "dates": [], "models": {}}
    if not year or not cruces_file.exists():
        return empty

    wb = load_workbook(cruces_file, read_only=True, data_only=True)
    if "Precios" not in wb.sheetnames:
        return empty

    segment_by_model = {}
    if "Segmento" in wb.sheetnames:
        ws_segment = wb["Segmento"]
        header = [cell.value for cell in next(ws_segment.iter_rows(min_row=1, max_row=1))]
        header_map = {name: index for index, name in enumerate(header)}
        model_index = header_map.get("MARCAMODELO")
        segment_index = header_map.get("Segmento")
        if model_index is not None and segment_index is not None:
            for row in ws_segment.iter_rows(min_row=2, values_only=True):
                model = row[model_index]
                segment = row[segment_index]
                if model:
                    segment_by_model[str(model).strip()] = str(segment or "").strip()

    ws_prices = wb["Precios"]
    header = [cell.value for cell in next(ws_prices.iter_rows(min_row=1, max_row=1))]
    header_map = {name: index for index, name in enumerate(header)}
    required = ["Fecha", "MARCAMODELO", "PRECIO"]
    if any(name not in header_map for name in required):
        return empty

    raw_prices: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for row in ws_prices.iter_rows(min_row=2, values_only=True):
        model = row[header_map["MARCAMODELO"]]
        price = row[header_map["PRECIO"]]
        price_date = parse_price_date(row[header_map["Fecha"]])
        if not model or price is None or not price_date or not price_date.startswith(year):
            continue
        try:
            numeric_price = float(price)
        except (TypeError, ValueError):
            continue
        raw_prices[str(model).strip()].append((price_date, numeric_price))

    models = {}
    all_dates = set()
    for model, items in sorted(raw_prices.items()):
        last_date = None
        last_price = None
        points = []
        for price_date, price in sorted(items):
            if price_date == last_date:
                continue
            if last_price is None or round(price, 4) != round(last_price, 4):
                points.append({"d": price_date, "p": price})
                all_dates.add(price_date)
            last_date = price_date
            last_price = price
        if points:
            models[model] = {
                "segment": segment_by_model.get(model, ""),
                "points": points,
            }

    return {
        "year": year,
        "source": str(cruces_file),
        "dates": sorted(all_dates),
        "models": models,
    }


def accessory_price_history(data: dict) -> dict:
    year = latest_year(data)
    sources = []
    for source in data.get("accessory_price_sources", []):
        path = Path(source.get("file", ""))
        source_date = str(source.get("date", ""))
        if not path.exists() or not source_date.startswith(year):
            continue
        candidates = extract_price_candidates(path)
        if candidates:
            sources.append((source_date, path, candidates))

    models = {}
    model_names = sorted({
        str(row.get("MARCAMODELO", "")).strip()
        for row in data.get("dashboard_data", [])
        if row.get("MARCAMODELO")
    })
    for model in model_names:
        points = []
        last_price = None
        for source_date, _path, candidates in sources:
            match = best_price_match(model, candidates)
            if not match:
                continue
            price = float(match["price"])
            if last_price is None or round(price, 4) != round(last_price, 4):
                points.append({"d": source_date, "p": price})
            last_price = price
        if points:
            models[model] = {"segment": "", "points": points}

    return {
        "year": year,
        "source": ", ".join(path.name for _date, path, _candidates in sources),
        "dates": [source_date for source_date, _path, _candidates in sources],
        "models": models,
    }


def build(input_json: Path, price_changes_json: Path | None = None) -> dict:
    data = json.loads(input_json.read_text(encoding="utf-8"))
    report_kind = data.get("report_kind", "mobiles")
    supervisor_by_pdv = load_supervisor_mapping()
    price_changes = {}
    if price_changes_json and price_changes_json.exists():
        price_changes = json.loads(price_changes_json.read_text(encoding="utf-8"))
    rows = [
        {
            "c": dashboard_channel(row["CANAL_AGRUPADO"]),
            "rg": row["REGIONAL_HONOR"],
            "sup": supervisor_for_pdv(row["PUNTODEVENTA"], supervisor_by_pdv),
            "p": row["PUNTODEVENTA"],
            "s": row["SEGMENTO_HONOR"],
            **(
                {"t": row.get("TIPO"), "st": row.get("SUBTIPO")}
                if report_kind == "accessories"
                else {}
            ),
            "b": row["MARCA"],
            "m": row["MARCAMODELO"],
            "w": row["SEMANA"],
            "f": row["FECHA"],
            "pe": row.get("PERIODO"),
            "y": row.get("ANIO"),
            "ml": row.get("MES_LABEL"),
            "d": row.get("DIA") or (row["FECHA"][-2:] if row["FECHA"] != "Sin fecha" else ""),
            "v": row["VENTAS"],
            "pr": row.get("PRECIO"),
        }
        for row in data["dashboard_data"]
    ]
    stock_rows = [
        {
            "c": dashboard_channel(row["CANAL_AGRUPADO"]),
            "rg": row["REGIONAL_HONOR"],
            "sup": supervisor_for_pdv(row["PUNTODEVENTA"], supervisor_by_pdv),
            "p": row["PUNTODEVENTA"],
            "s": row["SEGMENTO_HONOR"],
            **(
                {"t": row.get("TIPO"), "st": row.get("SUBTIPO")}
                if report_kind == "accessories"
                else {}
            ),
            "b": row["MARCA"],
            "m": row["MARCAMODELO"],
            "f": row["FECHA"],
            "pe": row.get("PERIODO"),
            "y": row.get("ANIO"),
            "ml": row.get("MES_LABEL"),
            "d": row.get("DIA") or (row["FECHA"][-2:] if row["FECHA"] != "Sin fecha" else ""),
            "q": row["STOCK"],
        }
        for row in data.get("stock_data", [])
    ]
    lists = dict(data["lists"])
    lists["years"] = [str(year) for year in lists.get("years") or data.get("years", [])]
    lists["periods"] = [period["PERIODO"] for period in data.get("periods", []) if period.get("PERIODO")]
    lists["period_labels"] = {**period_label_map(data.get("periods", [])), **lists.get("period_labels", {})}
    lists["channels"] = list(dict.fromkeys(dashboard_channel(channel) for channel in lists.get("channels", [])))
    lists["supervisors"] = sorted(
        {row.get("sup", "Otros") for row in [*rows, *stock_rows] if row.get("sup")},
        key=lambda value: value.casefold(),
    )
    return {
        "meta": {
            "report_kind": report_kind,
            "total_sales": data["total_sales"],
            "filtered_rows": data["filtered_rows"],
            "stock_total": data.get("stock_total", 0),
            "stock_filtered_rows": data.get("stock_filtered_rows", 0),
            "source_file": data["source_file"],
            "stock_source_file": data.get("stock_source_file", ""),
            "stock_file_summary": data.get("stock_file_summary", {}),
            "cruces_file": data["cruces_file"],
            "generated_at": data["generated_at"],
            "weeks": data["week_ranges"],
            "periods": data.get("periods", []),
            "years": data.get("years", []),
            "dates": data["dates"],
            "stock_dates": data.get("stock_dates", []),
            "cross_summary": data.get("cross_summary", {}),
            "stock_note": data.get("stock_note", ""),
        },
        "lists": lists,
        "rows": rows,
        "stockRows": stock_rows,
        "modelPrices": weighted_model_prices(data["dashboard_data"]),
        "priceChanges": price_changes,
        "priceHistory": (
            accessory_price_history(data)
            if report_kind == "accessories"
            else price_history_from_cruces(data)
        ),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Convierte el JSON agregado al formato compacto del dashboard web.")
    parser.add_argument("input_json")
    parser.add_argument("output_js")
    parser.add_argument("--price-changes-json")
    args = parser.parse_args()
    payload = build(Path(args.input_json), Path(args.price_changes_json) if args.price_changes_json else None)
    out = Path(args.output_js)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("window.REPORT_DATA = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    print(json.dumps({"rows": len(payload["rows"]), "total_sales": payload["meta"]["total_sales"]}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
