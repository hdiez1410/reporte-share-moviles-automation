from __future__ import annotations

import argparse
import json
import math
import re
import struct
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from extract_sales_report_data import (
    DETAIL_FIELDS,
    FILTERS,
    MONTH_ABBR,
    SHEET_MEMBER,
    SHARED_STRINGS_MEMBER,
    add_table_sale,
    build_share_rows,
    date_label,
    decode_rk,
    group_channel,
    iter_records,
    load_crosses,
    lookup_price,
    make_table_accumulator,
    norm_key,
    normalized,
    parse_shared_strings,
    pct,
    sort_key,
    transaction_date,
    yyyymmdd,
)

STOCK_SHEET_MEMBER = "xl/worksheets/sheet1.bin"  # workbook order -> Base Stocks
STOCK_SHEET_NAME = "Base Stocks"
STOCK_USO_VALUES = {"Normal", "Pack"}
STOCK_DETAIL_FIELDS = [
    "PERIODO",
    "DIA",
    "CANAL",
    "PUNTODEVENTA",
    "MARCAMODELO",
    "MARCA",
    "FAMILIA",
    "TIPO",
    "ESTADO",
    "USO",
    "ITEMS",
]

SPANISH_MONTHS = {
    "enero": "01",
    "febrero": "02",
    "marzo": "03",
    "abril": "04",
    "mayo": "05",
    "junio": "06",
    "julio": "07",
    "agosto": "08",
    "setiembre": "09",
    "septiembre": "09",
    "octubre": "10",
    "noviembre": "11",
    "diciembre": "12",
}


def month_label(period: str) -> str:
    if len(period) != 6 or not period.isdigit():
        return "Sin periodo"
    year = int(period[:4])
    month = int(period[4:6])
    return f"{MONTH_ABBR.get(month, str(month))}-{str(year)[2:]}"


def week_from_date_unique(date_value):
    if date_value is None:
        return "Sin fecha", "", None, None
    iso = date_value.isocalendar()
    week_id = f"{iso.year}-W{iso.week:02d}"
    monday = date_value - timedelta(days=date_value.weekday())
    sunday = monday + timedelta(days=6)
    return week_id, str(iso.week), monday, sunday


def make_accumulator() -> dict:
    return {
        "source_rows": 0,
        "filtered_rows": 0,
        "stock_source_rows": 0,
        "stock_filtered_rows": 0,
        "stock_total": 0.0,
        "total_sales": 0.0,
        "signed_items_total": 0.0,
        "weeks": defaultdict(float),
        "week_meta": {},
        "periods": defaultdict(float),
        "period_meta": {},
        "brands": defaultdict(float),
        "week_brand": defaultdict(float),
        "pdv_names": set(),
        "models": set(),
        "regionals": set(),
        "segmentos_honor": set(),
        "years": set(),
        "price_match_counts": Counter(),
        "missing_crosses": Counter(),
        "missing_regional_pdvs": Counter(),
        "missing_segment_models": Counter(),
        "missing_price_models": Counter(),
        "value_presence": defaultdict(Counter),
        "dashboard_sales": defaultdict(float),
        "dashboard_stock": defaultdict(float),
        "dates": set(),
        "stock_dates": set(),
        "stock_source_file": "",
        "stock_file_summary": {},
        "source_files": [],
        "source_file_summaries": [],
        "tables": {
            "channel": make_table_accumulator(),
            "pdv": make_table_accumulator(),
            "segment": make_table_accumulator(),
            "subsegment": make_table_accumulator(),
            "segmento_honor": make_table_accumulator(),
        },
    }


def required_columns(header: list[str]) -> dict[str, int]:
    col_idx = {normalized(name): i for i, name in enumerate(header) if normalized(name)}
    missing = [field for field in [*FILTERS, "ITEMS", *DETAIL_FIELDS] if field not in col_idx]
    if missing:
        raise RuntimeError(f"Columnas no encontradas: {', '.join(missing)}")
    return col_idx


def required_stock_columns(header: list[str]) -> dict[str, int]:
    col_idx = {normalized(name): i for i, name in enumerate(header) if normalized(name)}
    missing = [field for field in STOCK_DETAIL_FIELDS if field not in col_idx]
    if missing:
        raise RuntimeError(f"Columnas de stock no encontradas: {', '.join(missing)}")
    return col_idx


def numeric_value(value):
    if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
        return float(value)
    text = normalized(value)
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def process_stock_record(record: dict[str, object], acc: dict, crosses: dict) -> None:
    acc["stock_source_rows"] += 1
    if normalized(record.get("ESTADO")) != "Nuevos":
        return
    if normalized(record.get("USO")) not in STOCK_USO_VALUES:
        return
    # El objetivo es stock de equipos moviles; sin este guardrail entran SIMs/accesorios.
    if normalized(record.get("FAMILIA")) != "Moviles":
        return
    if normalized(record.get("TIPO")) != "Celular":
        return

    items = numeric_value(record.get("ITEMS"))
    if items is None or abs(items) < 1e-12:
        return

    date_value = transaction_date(record.get("PERIODO"), record.get("DIA"))
    date_label_value = date_value.isoformat() if date_value else "Sin fecha"
    period = normalized(record.get("PERIODO"))
    year = period[:4] if len(period) >= 4 and period[:4].isdigit() else (str(date_value.year) if date_value else "Sin año")
    period_value = period if len(period) == 6 and period.isdigit() else "Sin periodo"
    original_channel = normalized(record.get("CANAL")) or "Sin canal"
    channel = group_channel(original_channel)
    pdv = normalized(record.get("PUNTODEVENTA")) or "Sin punto de venta"
    model = normalized(record.get("MARCAMODELO")) or "Sin modelo"
    brand = normalized(record.get("MARCA")) or "Sin marca"
    regional = crosses["regional_map"].get(norm_key(pdv), "Sin regional")
    segmento_honor = crosses["segmento_map"].get(norm_key(model), "Sin segmento Honor")

    key = (channel, regional, pdv, segmento_honor, brand, model, date_label_value, period_value, year)
    acc["dashboard_stock"][key] += items
    acc["stock_filtered_rows"] += 1
    acc["stock_total"] += items
    if date_value:
        acc["stock_dates"].add(date_value)


def process_record(record: dict[str, object], acc: dict, crosses: dict) -> None:
    acc["source_rows"] += 1
    if any(normalized(record.get(field)) != expected for field, expected in FILTERS.items()):
        return

    items = numeric_value(record.get("ITEMS"))
    if items is None:
        return

    sales = -items
    if abs(sales) < 1e-12:
        return

    date_value = transaction_date(record.get("PERIODO"), record.get("DIA"))
    date_key = yyyymmdd(date_value)
    week, week_number, week_start, week_end = week_from_date_unique(date_value)
    period = normalized(record.get("PERIODO"))
    year = period[:4] if len(period) >= 4 and period[:4].isdigit() else (str(date_value.year) if date_value else "Sin año")
    period_value = period if len(period) == 6 and period.isdigit() else "Sin periodo"
    brand = normalized(record.get("MARCA")) or "Sin marca"
    original_channel = normalized(record.get("CANAL")) or "Sin canal"
    channel = group_channel(original_channel)
    pdv = normalized(record.get("PUNTODEVENTA")) or "Sin punto de venta"
    gama = normalized(record.get("GAMA")) or "Sin gama"
    subgama = normalized(record.get("SUBGAMA")) or "Sin subgama"
    model = normalized(record.get("MARCAMODELO")) or "Sin modelo"

    regional = crosses["regional_map"].get(norm_key(pdv), "Sin regional")
    segmento_honor = crosses["segmento_map"].get(norm_key(model), "Sin segmento Honor")
    price, price_status = lookup_price(crosses, model, date_key)

    acc["filtered_rows"] += 1
    acc["total_sales"] += sales
    acc["signed_items_total"] += items
    acc["weeks"][week] += sales
    if week_start and week not in acc["week_meta"]:
        acc["week_meta"][week] = (week_number, week_start, week_end)
    acc["periods"][period_value] += sales
    if period_value != "Sin periodo":
        acc["period_meta"][period_value] = {"PERIODO": period_value, "ANIO": year, "LABEL": month_label(period_value)}
    acc["years"].add(year)
    acc["brands"][brand] += sales
    acc["week_brand"][(week, brand)] += sales
    acc["pdv_names"].add(pdv)
    acc["models"].add(model)
    acc["regionals"].add(regional)
    acc["segmentos_honor"].add(segmento_honor)
    acc["value_presence"]["CANAL_ORIGINAL"][original_channel] += 1
    acc["value_presence"]["CANAL_AGRUPADO"][channel] += 1
    acc["value_presence"]["GAMA"][gama] += 1
    acc["value_presence"]["SUBGAMA"][subgama] += 1
    acc["value_presence"]["REGIONAL_HONOR"][regional] += 1
    acc["value_presence"]["SEGMENTO_HONOR"][segmento_honor] += 1
    acc["price_match_counts"][price_status] += 1
    if regional == "Sin regional":
        acc["missing_crosses"]["REGIONAL"] += sales
        acc["missing_regional_pdvs"][pdv] += sales
    if segmento_honor == "Sin segmento Honor":
        acc["missing_crosses"]["SEGMENTO_HONOR"] += sales
        acc["missing_segment_models"][model] += sales
    if price is None:
        acc["missing_crosses"]["PRECIO"] += sales
        acc["missing_price_models"][model] += sales

    date_label_value = date_value.isoformat() if date_value else "Sin fecha"
    key = (channel, regional, pdv, segmento_honor, brand, model, week, date_label_value, period_value, year)
    acc["dashboard_sales"][key] += sales
    if date_value:
        acc["dates"].add(date_value)

    add_table_sale(acc["tables"]["channel"], (channel, regional, brand), (channel, regional), (channel,), week, sales, None)
    add_table_sale(acc["tables"]["pdv"], (channel, regional, pdv, brand, model), (channel, pdv), (channel, pdv), week, sales, price)
    add_table_sale(acc["tables"]["segment"], (gama, regional, brand, model), (gama, regional), (gama,), week, sales, price)
    add_table_sale(acc["tables"]["subsegment"], (subgama, regional, brand, model), (subgama, regional), (subgama,), week, sales, price)
    add_table_sale(acc["tables"]["segmento_honor"], (segmento_honor, regional, brand, model), (segmento_honor, regional), (segmento_honor,), week, sales, price)


def process_xlsx(path: Path, acc: dict, crosses: dict, include_sales: bool = True, include_stock: bool = True) -> dict:
    before_source = acc["source_rows"]
    before_filtered = acc["filtered_rows"]
    before_stock_source = acc["stock_source_rows"]
    before_stock_filtered = acc["stock_filtered_rows"]
    wb = load_workbook(path, read_only=True, data_only=True)
    if include_sales and "Base Movs" not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"{path.name} no tiene la hoja Base Movs.")
    stock_sheet_found = STOCK_SHEET_NAME in wb.sheetnames
    if include_sales:
        ws = wb["Base Movs"]
        rows = ws.iter_rows(values_only=True)
        header = [normalized(value) for value in next(rows)]
        col_idx = required_columns(header)
        for row in rows:
            record = {field: row[index] if index < len(row) else None for field, index in col_idx.items()}
            process_record(record, acc, crosses)
    if include_stock and stock_sheet_found:
        ws_stock = wb[STOCK_SHEET_NAME]
        stock_rows = ws_stock.iter_rows(values_only=True)
        stock_header = [normalized(value) for value in next(stock_rows)]
        stock_col_idx = required_stock_columns(stock_header)
        for row in stock_rows:
            record = {field: row[index] if index < len(row) else None for field, index in stock_col_idx.items()}
            process_stock_record(record, acc, crosses)
    wb.close()
    return {
        "file": str(path),
        "type": "xlsx",
        "source_rows": acc["source_rows"] - before_source,
        "filtered_rows": acc["filtered_rows"] - before_filtered,
        "stock_source_rows": acc["stock_source_rows"] - before_stock_source,
        "stock_filtered_rows": acc["stock_filtered_rows"] - before_stock_filtered,
        "stock_sheet_found": stock_sheet_found,
    }


def xlsb_header(data: bytes, shared: list[str]) -> list[str]:
    row_values: dict[int, object] = {}
    current_row = None
    for record_type, payload in iter_records(data):
        if record_type == 0:
            row_idx = int.from_bytes(payload[0:4], "little", signed=False)
            if current_row is not None and row_idx != current_row:
                break
            current_row = row_idx
        elif current_row is not None and record_type == 7 and len(payload) >= 12:
            col = int.from_bytes(payload[0:4], "little", signed=False)
            sst_idx = int.from_bytes(payload[8:12], "little", signed=False)
            row_values[col] = shared[sst_idx] if 0 <= sst_idx < len(shared) else ""
        elif current_row is not None and record_type == 5 and len(payload) >= 16:
            col = int.from_bytes(payload[0:4], "little", signed=False)
            row_values[col] = struct.unpack("<d", payload[8:16])[0]
        elif current_row is not None and record_type == 2 and len(payload) >= 12:
            col = int.from_bytes(payload[0:4], "little", signed=False)
            row_values[col] = decode_rk(payload[8:12])
    return [normalized(row_values.get(i)) for i in range(max(row_values.keys(), default=-1) + 1)]


def identify_xlsb_sheets(zf: zipfile.ZipFile, shared: list[str]) -> tuple[bytes | None, bytes | None]:
    worksheet_members = sorted(
        member
        for member in zf.namelist()
        if re.fullmatch(r"xl/worksheets/sheet\d+\.bin", member)
    )
    sales_data = None
    stock_data = None
    sales_required = {*FILTERS, "ITEMS", *DETAIL_FIELDS}
    stock_required = set(STOCK_DETAIL_FIELDS)
    for member in worksheet_members:
        data = zf.read(member)
        fields = {normalized(value) for value in xlsb_header(data, shared) if normalized(value)}
        if sales_data is None and sales_required.issubset(fields):
            sales_data = data
        elif stock_data is None and stock_required.issubset(fields):
            stock_data = data
    return sales_data, stock_data


def process_xlsb(path: Path, acc: dict, crosses: dict, include_sales: bool = True, include_stock: bool = True) -> dict:
    before_source = acc["source_rows"]
    before_filtered = acc["filtered_rows"]
    before_stock_source = acc["stock_source_rows"]
    before_stock_filtered = acc["stock_filtered_rows"]
    with zipfile.ZipFile(path) as zf:
        shared = parse_shared_strings(zf)
        sales_sheet_data, stock_sheet_data = identify_xlsb_sheets(zf, shared)
        data = sales_sheet_data if include_sales else None
        stock_data = stock_sheet_data if include_stock else None
    if include_sales and data is None:
        raise RuntimeError(f"{path.name} no tiene una hoja Base Movs reconocible.")

    header: list[str] = []
    col_idx: dict[str, int] = {}
    current_row = None
    row_values: dict[int, object] = {}

    def finish_row(row_idx, values):
        nonlocal header, col_idx
        if row_idx is None:
            return
        if row_idx == 0:
            header = [normalized(values.get(i)) for i in range(max(values.keys(), default=-1) + 1)]
            col_idx = required_columns(header)
            return
        record = {field: values.get(index) for field, index in col_idx.items()}
        process_record(record, acc, crosses)

    if data is not None:
        for record_type, payload in iter_records(data):
            if record_type == 0:
                finish_row(current_row, row_values)
                current_row = int.from_bytes(payload[0:4], "little", signed=False)
                row_values = {}
            elif current_row is not None and record_type == 7 and len(payload) >= 12:
                col = int.from_bytes(payload[0:4], "little", signed=False)
                sst_idx = int.from_bytes(payload[8:12], "little", signed=False)
                row_values[col] = shared[sst_idx] if 0 <= sst_idx < len(shared) else ""
            elif current_row is not None and record_type == 5 and len(payload) >= 16:
                col = int.from_bytes(payload[0:4], "little", signed=False)
                row_values[col] = struct.unpack("<d", payload[8:16])[0]
            elif current_row is not None and record_type == 2 and len(payload) >= 12:
                col = int.from_bytes(payload[0:4], "little", signed=False)
                row_values[col] = decode_rk(payload[8:12])
        finish_row(current_row, row_values)

    stock_sheet_found = stock_data is not None
    if stock_data is not None:
        stock_header: list[str] = []
        stock_col_idx: dict[str, int] = {}
        stock_current_row = None
        stock_row_values: dict[int, object] = {}

        def finish_stock_row(row_idx, values):
            nonlocal stock_header, stock_col_idx
            if row_idx is None:
                return
            if row_idx == 0:
                stock_header = [normalized(values.get(i)) for i in range(max(values.keys(), default=-1) + 1)]
                stock_col_idx = required_stock_columns(stock_header)
                return
            record = {field: values.get(index) for field, index in stock_col_idx.items()}
            process_stock_record(record, acc, crosses)

        for record_type, payload in iter_records(stock_data):
            if record_type == 0:
                finish_stock_row(stock_current_row, stock_row_values)
                stock_current_row = int.from_bytes(payload[0:4], "little", signed=False)
                stock_row_values = {}
            elif stock_current_row is not None and record_type == 7 and len(payload) >= 12:
                col = int.from_bytes(payload[0:4], "little", signed=False)
                sst_idx = int.from_bytes(payload[8:12], "little", signed=False)
                stock_row_values[col] = shared[sst_idx] if 0 <= sst_idx < len(shared) else ""
            elif stock_current_row is not None and record_type == 5 and len(payload) >= 16:
                col = int.from_bytes(payload[0:4], "little", signed=False)
                stock_row_values[col] = struct.unpack("<d", payload[8:16])[0]
            elif stock_current_row is not None and record_type == 2 and len(payload) >= 12:
                col = int.from_bytes(payload[0:4], "little", signed=False)
                stock_row_values[col] = decode_rk(payload[8:12])
        finish_stock_row(stock_current_row, stock_row_values)

    return {
        "file": str(path),
        "type": "xlsb",
        "source_rows": acc["source_rows"] - before_source,
        "filtered_rows": acc["filtered_rows"] - before_filtered,
        "stock_source_rows": acc["stock_source_rows"] - before_stock_source,
        "stock_filtered_rows": acc["stock_filtered_rows"] - before_stock_filtered,
        "stock_sheet_found": stock_sheet_found,
    }


def period_sort_key(value: str):
    return int(value) if value.isdigit() else 0


def period_from_filename(path_text: str) -> str | None:
    name = Path(path_text).name.casefold()
    numeric = re.search(r"(20\d{2})(0[1-9]|1[0-2])(?:\d{2})?", name)
    if numeric:
        return f"{numeric.group(1)}{numeric.group(2)}"
    year = re.search(r"(20\d{2})", name)
    if not year:
        return None
    for month_name, month_number in SPANISH_MONTHS.items():
        if month_name in name:
            return f"{year.group(1)}{month_number}"
    return None


def stock_source_sort_key(path: Path):
    period = period_from_filename(str(path)) or ""
    return (period_sort_key(period), path.name.casefold())


def source_choice_key(path: Path) -> tuple[int, int, int, float, str]:
    name = path.name.casefold()
    period = period_from_filename(path.name) or "000000"
    exact_date = re.search(r"(20\d{6})", name)
    date_score = int(exact_date.group(1)) if exact_date else int(f"{period}00")
    corrected_april = 1 if "base movs" in name and "abril" in name and "2026" in name else 0
    if corrected_april:
        date_score = 20260499
    cierre = 1 if "cierre" in name else 0
    return (date_score, corrected_april, cierre, path.stat().st_mtime, path.name)


def select_sources_by_period(files: list[Path]) -> list[Path]:
    selected: dict[str, Path] = {}
    for path in files:
        period = period_from_filename(path.name)
        if not period:
            print(json.dumps({"warning": "sin_periodo", "file": path.name}, ensure_ascii=False), flush=True)
            continue
        previous = selected.get(period)
        if previous is None or source_choice_key(path) >= source_choice_key(previous):
            selected[period] = path
    return [selected[period] for period in sorted(selected, key=period_sort_key)]


def build_output(acc: dict, crosses: dict, source_dir: Path) -> dict:
    weeks = sorted(acc["weeks"].keys(), key=lambda value: acc["week_meta"].get(value, ("", datetime.max.date(), None))[1])
    brands = sorted(acc["brands"].keys(), key=lambda b: (-acc["brands"][b], b))
    sales_matrix = []
    share_matrix = []
    for brand in brands:
        sales_row = {"MARCA": brand, "TOTAL": round(acc["brands"][brand], 6), "SHARE_TOTAL": pct(acc["brands"][brand], acc["total_sales"])}
        share_row = {"MARCA": brand, "SHARE_TOTAL": pct(acc["brands"][brand], acc["total_sales"])}
        for week in weeks:
            sales = acc["week_brand"].get((week, brand), 0.0)
            sales_row[week] = round(sales, 6)
            share_row[week] = pct(sales, acc["weeks"][week])
        sales_matrix.append(sales_row)
        share_matrix.append(share_row)

    week_ranges = []
    for week in weeks:
        week_number, start, end = acc["week_meta"].get(week, ("", None, None))
        year = str(start.isocalendar().year) if start else ""
        week_ranges.append(
            {
                "SEMANA": week,
                "NUMERO": week_number,
                "ANIO": year,
                "DESDE": start.isoformat() if start else "",
                "HASTA": end.isoformat() if end else "",
                "LABEL": f"{year} S{week_number} ({date_label(start)} al {date_label(end)})" if start and end else week,
                "SHORT": f"S{week_number}" if week_number else week,
            }
        )

    periods = [
        {
            "PERIODO": period,
            "ANIO": meta["ANIO"],
            "LABEL": meta["LABEL"],
            "VENTAS": round(acc["periods"][period], 6),
        }
        for period, meta in sorted(acc["period_meta"].items(), key=lambda item: period_sort_key(item[0]))
    ]
    weekly_totals = [{"SEMANA": w, "VENTAS": round(acc["weeks"][w], 6)} for w in weeks]
    brand_totals = [{"MARCA": b, "VENTAS": round(acc["brands"][b], 6), "SHARE": pct(acc["brands"][b], acc["total_sales"])} for b in brands]

    share_by_channel = build_share_rows(acc["tables"]["channel"], weeks, ["CANAL_AGRUPADO", "REGIONAL_HONOR", "MARCA"], "TOTAL_CANAL", "TOTAL_CANAL_REGIONAL")
    share_by_pdv = build_share_rows(acc["tables"]["pdv"], weeks, ["CANAL_AGRUPADO", "REGIONAL_HONOR", "PUNTODEVENTA", "MARCA", "MARCAMODELO"], "TOTAL_PDV", None, True)
    share_by_segment = build_share_rows(acc["tables"]["segment"], weeks, ["SEGMENTO_GAMA", "REGIONAL_HONOR", "MARCA", "MARCAMODELO"], "TOTAL_SEGMENTO", "TOTAL_SEGMENTO_REGIONAL", True)
    share_by_subsegment = build_share_rows(acc["tables"]["subsegment"], weeks, ["SUBGAMA", "REGIONAL_HONOR", "MARCA", "MARCAMODELO"], "TOTAL_SUBGAMA", "TOTAL_SUBGAMA_REGIONAL", True)
    share_by_segmento_honor = build_share_rows(acc["tables"]["segmento_honor"], weeks, ["SEGMENTO_HONOR", "REGIONAL_HONOR", "MARCA", "MARCAMODELO"], "TOTAL_SEGMENTO_HONOR", "TOTAL_SEGMENTO_HONOR_REGIONAL", True)

    channel_totals_for_order = {
        channel: acc["tables"]["channel"]["primary_sales"][(channel,)]
        for channel in {row["CANAL_AGRUPADO"] for row in share_by_channel}
    }
    share_by_pdv.sort(
        key=lambda row: (
            -channel_totals_for_order.get(row["CANAL_AGRUPADO"], 0.0),
            -row["TOTAL_PDV"],
            -row["VENTAS_MES"],
            row["PUNTODEVENTA"],
            row["MARCA"],
            row["MARCAMODELO"],
        )
    )

    dashboard_data = [
        {
            "CANAL_AGRUPADO": channel,
            "REGIONAL_HONOR": regional,
            "PUNTODEVENTA": pdv,
            "SEGMENTO_HONOR": segment,
            "MARCA": brand,
            "MARCAMODELO": model,
            "SEMANA": week,
            "FECHA": date_value,
            "DIA": date_value[-2:] if date_value != "Sin fecha" else "",
            "PERIODO": period,
            "ANIO": year,
            "MES_LABEL": month_label(period),
            "VENTAS": round(sales, 6),
            "PRECIO": lookup_price(crosses, model, yyyymmdd(datetime.strptime(date_value, "%Y-%m-%d").date()) if date_value != "Sin fecha" else None)[0],
        }
        for (channel, regional, pdv, segment, brand, model, week, date_value, period, year), sales in acc["dashboard_sales"].items()
    ]
    dashboard_data.sort(
        key=lambda row: (
            row["ANIO"],
            row["PERIODO"],
            row["CANAL_AGRUPADO"],
            row["REGIONAL_HONOR"],
            row["PUNTODEVENTA"],
            row["SEGMENTO_HONOR"],
            row["MARCA"],
            row["MARCAMODELO"],
            row["SEMANA"],
            row["FECHA"],
        )
    )

    stock_data = [
        {
            "CANAL_AGRUPADO": channel,
            "REGIONAL_HONOR": regional,
            "PUNTODEVENTA": pdv,
            "SEGMENTO_HONOR": segment,
            "MARCA": brand,
            "MARCAMODELO": model,
            "FECHA": date_value,
            "DIA": date_value[-2:] if date_value != "Sin fecha" else "",
            "PERIODO": period,
            "ANIO": year,
            "MES_LABEL": month_label(period),
            "STOCK": round(stock, 6),
        }
        for (channel, regional, pdv, segment, brand, model, date_value, period, year), stock in acc["dashboard_stock"].items()
    ]
    stock_data.sort(
        key=lambda row: (
            row["ANIO"],
            row["PERIODO"],
            row["CANAL_AGRUPADO"],
            row["REGIONAL_HONOR"],
            row["PUNTODEVENTA"],
            row["SEGMENTO_HONOR"],
            row["MARCA"],
            row["MARCAMODELO"],
            row["FECHA"],
        )
    )

    unique_models = sorted(acc["models"])
    unique_pdvs = sorted(acc["pdv_names"])
    unique_regionals = sorted(acc["regionals"])
    unique_segmentos_honor = sorted(acc["segmentos_honor"])
    years = sorted(acc["years"])

    return {
        "source_file": str(source_dir),
        "source_files": acc["source_files"],
        "source_file_summaries": acc["source_file_summaries"],
        "stock_source_file": acc.get("stock_source_file", ""),
        "stock_file_summary": acc.get("stock_file_summary", {}),
        "source_sheet": "Base Movs",
        "cruces_file": crosses["source"],
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "filters": FILTERS,
        "metric_note": "VENTAS = -SUM(ITEMS), porque los movimientos de venta figuran como salidas negativas de stock.",
        "week_note": "Semana calculada desde PERIODO + DIA con corte lunes a domingo; no se usa el campo SEMANA de la base.",
        "channel_note": "Canales agrupados: TPF/TF/Tiendas Propias Franquicias=Tiendas; Empresas=Empresas; Retail/Islas=Retail&Islas; OMNIX/Logixtal/Postventa/Brightstar/Brighstar/Televenta=Remotos; TIENDAS EXPRESS/TEX/TEX SATELITE=Tiendas Express; Central=Central.",
        "segment_note": "Se usa GAMA como segmento principal, SUBGAMA como subsegmento y CRUCES/Segmento como Segmento Honor.",
        "crosses_note": "Regional Honor cruza PUNTODEVENTA contra CRUCES/REGIONAL. Segmento Honor cruza MARCAMODELO contra CRUCES/Segmento.",
        "stock_note": "STOCK actual = SUM(ITEMS) desde Base Stocks de la última base Oracle disponible, con ESTADO=Nuevos, USO en Normal/Pack y equipos moviles (FAMILIA=Moviles, TIPO=Celular). No se acumulan fotos historicas de stock.",
        "source_rows": acc["source_rows"],
        "filtered_rows": acc["filtered_rows"],
        "stock_source_rows": acc["stock_source_rows"],
        "stock_filtered_rows": acc["stock_filtered_rows"],
        "stock_total": round(acc["stock_total"], 6),
        "total_sales": round(acc["total_sales"], 6),
        "signed_items_total": round(acc["signed_items_total"], 6),
        "weeks": weeks,
        "week_ranges": week_ranges,
        "periods": periods,
        "years": years,
        "brands": brands,
        "weekly_totals": weekly_totals,
        "brand_totals": brand_totals,
        "weekly_sales_matrix": sales_matrix,
        "weekly_share_matrix": share_matrix,
        "dates": [
            {"FECHA": date.isoformat(), "DIA": f"{date.day:02d}", "PERIODO": f"{date.year}{date.month:02d}", "ANIO": str(date.year), "LABEL": f"{date.day:02d}-{MONTH_ABBR[date.month]}"}
            for date in sorted(acc["dates"])
        ],
        "dashboard_data": dashboard_data,
        "stock_data": stock_data,
        "stock_dates": [
            {"FECHA": date.isoformat(), "DIA": f"{date.day:02d}", "PERIODO": f"{date.year}{date.month:02d}", "ANIO": str(date.year), "LABEL": f"{date.day:02d}-{MONTH_ABBR[date.month]}"}
            for date in sorted(acc["stock_dates"])
        ],
        "lists": {
            "years": years,
            "periods": [item["PERIODO"] for item in periods],
            "period_labels": {item["PERIODO"]: item["LABEL"] for item in periods},
            "channels": sorted({row["CANAL_AGRUPADO"] for row in share_by_channel}, key=lambda c: -acc["tables"]["channel"]["primary_sales"][(c,)]),
            "regionals": unique_regionals,
            "pdvs": unique_pdvs,
            "segmentos_honor": unique_segmentos_honor,
            "models": unique_models,
            "brands": brands,
        },
        "channel_groups": sorted({row["CANAL_AGRUPADO"] for row in share_by_channel}, key=lambda c: -acc["tables"]["channel"]["primary_sales"][(c,)]),
        "share_by_channel": share_by_channel,
        "share_by_pdv": share_by_pdv,
        "share_by_segment": share_by_segment,
        "share_by_subsegment": share_by_subsegment,
        "share_by_segmento_honor": share_by_segmento_honor,
        "cross_summary": {
            "cruces_counts": crosses["counts"],
            "price_match_counts": dict(acc["price_match_counts"]),
            "missing_sales": {key: round(value, 6) for key, value in acc["missing_crosses"].items()},
            "missing_regional_pdvs": [{"PUNTODEVENTA": key, "VENTAS": round(value, 6)} for key, value in acc["missing_regional_pdvs"].most_common(50)],
            "missing_segment_models": [{"MARCAMODELO": key, "VENTAS": round(value, 6)} for key, value in acc["missing_segment_models"].most_common(50)],
            "missing_price_models": [{"MARCAMODELO": key, "VENTAS": round(value, 6)} for key, value in acc["missing_price_models"].most_common(50)],
        },
        "counts": {
            "weeks": len(weeks),
            "periods": len(periods),
            "years": len(years),
            "brands": len(brands),
            "channels": len({row["CANAL_AGRUPADO"] for row in share_by_channel}),
            "pdv": len(acc["pdv_names"]),
            "models": len(acc["models"]),
            "regionals": len(acc["regionals"]),
            "segments_gama": len({row["SEGMENTO_GAMA"] for row in share_by_segment}),
            "subsegments": len({row["SUBGAMA"] for row in share_by_subsegment}),
            "segmentos_honor": len(acc["segmentos_honor"]),
            "stock_rows": len(stock_data),
            "stock_dates": len(acc["stock_dates"]),
        },
    }


def extract_sources(source_dir: Path, cruces_path: Path) -> dict:
    crosses = load_crosses(cruces_path)
    acc = make_accumulator()
    candidates = sorted(
        path
        for path in source_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsb"} and not path.name.startswith("._")
    )
    files = select_sources_by_period(candidates)
    skipped = sorted(set(candidates) - set(files), key=lambda path: (period_from_filename(path.name) or "000000", path.name))
    for path in skipped:
        print(json.dumps({"skipped_duplicate_period": path.name, "period": period_from_filename(path.name)}, ensure_ascii=False), flush=True)
    for path in files:
        acc["source_files"].append(str(path))
        if path.suffix.lower() == ".xlsx":
            summary = process_xlsx(path, acc, crosses, include_stock=False)
        else:
            summary = process_xlsb(path, acc, crosses, include_stock=False)
        acc["source_file_summaries"].append(summary)
        print(json.dumps(summary, ensure_ascii=False), flush=True)
    if files:
        stock_path = max(files, key=stock_source_sort_key)
        if stock_path.suffix.lower() == ".xlsx":
            stock_summary = process_xlsx(stock_path, acc, crosses, include_sales=False, include_stock=True)
        else:
            stock_summary = process_xlsb(stock_path, acc, crosses, include_sales=False, include_stock=True)
        acc["stock_source_file"] = str(stock_path)
        acc["stock_file_summary"] = stock_summary
        print(json.dumps({"stock_snapshot": stock_summary}, ensure_ascii=False), flush=True)
    return build_output(acc, crosses, source_dir)


def main() -> int:
    parser = argparse.ArgumentParser(description="Extrae y agrega todas las bases en BD Oracle para el dashboard.")
    parser.add_argument("output")
    parser.add_argument("--source-dir", required=True)
    parser.add_argument("--cruces", required=True)
    args = parser.parse_args()

    data = extract_sources(Path(args.source_dir), Path(args.cruces))
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "filtered_rows": data["filtered_rows"],
                "total_sales": data["total_sales"],
                "counts": data["counts"],
                "cross_summary": data["cross_summary"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
