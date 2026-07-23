from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


TARGET_SHEET = "PORTA PAGO UNICO REGULAR"
TARGET_HEADER = "63.9 69.9 74.9 79.9"
MODEL_COL = 4
PRICE_COL = 7


def normalized(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def norm_key(value) -> str:
    return " ".join(normalized(value).casefold().split())


def parse_date_from_name(path: Path) -> date:
    match = re.search(r"(\d{2})-(\d{2})-(\d{4})", path.name)
    if not match:
        raise ValueError(f"No pude encontrar fecha en el nombre: {path.name}")
    day, month, year = map(int, match.groups())
    return date(year, month, day)


def yyyymmdd(value: date) -> int:
    return value.year * 10000 + value.month * 100 + value.day


def date_range(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def numeric_price(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalized(value)
    if not text or text in {"-", "NA", "N/A"}:
        return None
    text = text.replace("S/", "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def find_header_row(ws) -> int:
    for row_idx in range(1, min(ws.max_row, 80) + 1):
        model_header = normalized(ws.cell(row_idx, MODEL_COL).value).casefold()
        price_header = normalized(ws.cell(row_idx, PRICE_COL).value)
        if model_header == "modelo" and price_header == TARGET_HEADER:
            return row_idx
    raise RuntimeError(f"No encontré el encabezado esperado en {ws.title}.")


def extract_prices(path: Path) -> dict[str, dict]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    if TARGET_SHEET not in wb.sheetnames:
        raise RuntimeError(f"{path.name} no tiene la hoja {TARGET_SHEET}.")
    ws = wb[TARGET_SHEET]
    header_row = find_header_row(ws)
    prices: OrderedDict[str, dict] = OrderedDict()
    skipped_non_price = 0
    duplicate_rows = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        model = normalized(ws.cell(row_idx, MODEL_COL).value)
        price = numeric_price(ws.cell(row_idx, PRICE_COL).value)
        if not model or model.casefold() == "modelo":
            continue
        if price is None:
            skipped_non_price += 1
            continue
        key = norm_key(model)
        if key in prices:
            duplicate_rows += 1
        prices[key] = {"MARCAMODELO": model, "PRECIO": price}

    return {
        "file": path.name,
        "sheet": TARGET_SHEET,
        "header_row": header_row,
        "models": list(prices.values()),
        "skipped_non_price": skipped_non_price,
        "duplicate_rows": duplicate_rows,
    }


def build_effective_ranges(price_files: list[Path], max_existing_date: date, latest_end: date | None = None) -> list[dict]:
    starts = sorted((parse_date_from_name(path), path) for path in price_files)
    ranges = []
    for idx, (start, path) in enumerate(starts):
        if idx + 1 < len(starts):
            end = starts[idx + 1][0] - timedelta(days=1)
        elif latest_end is not None:
            # Keep the latest list date in the historical price table even when
            # sales data ends before that list starts. Future prices still will
            # not be used for past sales because lookups require price_date <= sale_date.
            end = max(latest_end, start)
        else:
            end = max(max_existing_date, start)
        ranges.append({"start": start, "end": end, "path": path})
    return ranges


def prices_rows_for_ranges(extractions: dict[str, dict], ranges: list[dict]) -> list[dict]:
    rows = []
    for item in ranges:
        extracted = extractions[item["path"].name]
        for day in date_range(item["start"], item["end"]):
            date_key = yyyymmdd(day)
            week = f"W{day.isocalendar().week}"
            for model_row in extracted["models"]:
                model = model_row["MARCAMODELO"]
                rows.append(
                    {
                        "Fecha": pd.Timestamp(day),
                        "FECHAFINALCOM": date_key,
                        "Week": week,
                        "MARCAMODELO": model,
                        "PRECIO": model_row["PRECIO"],
                        "MODELOFECHAFINAL": f"{model}{date_key}",
                    }
                )
    return rows


def write_precios_sheet(cruces_path: Path, output_path: Path, price_df: pd.DataFrame):
    wb = load_workbook(cruces_path)
    if "Precios" not in wb.sheetnames:
        raise RuntimeError("CRUCES no tiene hoja Precios.")
    idx = wb.sheetnames.index("Precios")
    ws = wb["Precios"]
    wb.remove(ws)
    ws = wb.create_sheet("Precios", idx)
    ws.append(list(price_df.columns))
    for row in price_df.itertuples(index=False, name=None):
        ws.append(list(row))
    ws.freeze_panes = "A2"
    widths = {
        "A": 13,
        "B": 14,
        "C": 9,
        "D": 36,
        "E": 12,
        "F": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Actualiza CRUCES/Precios con listas Entel por fecha efectiva.")
    parser.add_argument("--cruces", required=True)
    parser.add_argument("--prices-dir", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--audit", required=True)
    parser.add_argument("--latest-end", help="Fecha final yyyy-mm-dd para la lista mas reciente.")
    args = parser.parse_args()

    cruces_path = Path(args.cruces)
    prices_dir = Path(args.prices_dir)
    price_files = sorted(path for path in prices_dir.glob("*.xlsm") if not path.name.startswith("._"))
    if not price_files:
        raise RuntimeError(f"No encontré listas .xlsm en {prices_dir}")

    old_prices = pd.read_excel(cruces_path, sheet_name="Precios")
    old_prices["FECHAFINALCOM"] = pd.to_numeric(old_prices["FECHAFINALCOM"], errors="coerce").astype("Int64")
    max_existing_key = int(old_prices["FECHAFINALCOM"].max())
    max_existing_date = datetime.strptime(str(max_existing_key), "%Y%m%d").date()

    latest_end = datetime.strptime(args.latest_end, "%Y-%m-%d").date() if args.latest_end else None
    ranges = build_effective_ranges(price_files, max_existing_date, latest_end)
    first_replacement_key = yyyymmdd(ranges[0]["start"])
    extractions = {path.name: extract_prices(path) for path in price_files}
    new_rows = prices_rows_for_ranges(extractions, ranges)
    new_prices = pd.DataFrame(new_rows, columns=["Fecha", "FECHAFINALCOM", "Week", "MARCAMODELO", "PRECIO", "MODELOFECHAFINAL"])

    kept_old = old_prices[old_prices["FECHAFINALCOM"] < first_replacement_key].copy()
    final_prices = pd.concat([kept_old, new_prices], ignore_index=True)
    final_prices["FECHAFINALCOM"] = final_prices["FECHAFINALCOM"].astype(int)
    final_prices = final_prices.sort_values(["FECHAFINALCOM", "MARCAMODELO"], kind="stable")

    segmento = pd.read_excel(cruces_path, sheet_name="Segmento")
    segmento_models = {norm_key(value) for value in segmento.get("MARCAMODELO", []) if normalized(value)}
    price_models = OrderedDict()
    for extracted in extractions.values():
        for row in extracted["models"]:
            price_models[norm_key(row["MARCAMODELO"])] = row["MARCAMODELO"]

    write_precios_sheet(cruces_path, Path(args.output), final_prices)

    audit = {
        "cruces_source": str(cruces_path),
        "cruces_output": str(Path(args.output)),
        "target_sheet": TARGET_SHEET,
        "target_column": "G",
        "target_plan": TARGET_HEADER,
        "effective_ranges": [
            {
                "file": item["path"].name,
                "from": item["start"].isoformat(),
                "to": item["end"].isoformat(),
                "models_with_price": len(extractions[item["path"].name]["models"]),
            }
            for item in ranges
        ],
        "old_price_rows": int(len(old_prices)),
        "kept_old_rows_before_first_list": int(len(kept_old)),
        "new_price_rows": int(len(new_prices)),
        "final_price_rows": int(len(final_prices)),
        "price_list_models": int(len(price_models)),
        "price_models_missing_in_segmento": sorted(
            [model for key, model in price_models.items() if key not in segmento_models],
            key=str.casefold,
        ),
        "extractions": {
            name: {
                "header_row": item["header_row"],
                "models_with_price": len(item["models"]),
                "skipped_non_price": item["skipped_non_price"],
                "duplicate_rows": item["duplicate_rows"],
            }
            for name, item in extractions.items()
        },
    }
    audit_path = Path(args.audit)
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
