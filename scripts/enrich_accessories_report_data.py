from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

from add_regional_overrides import REGIONAL_OVERRIDES
from extract_sales_report_data import group_channel, norm_key, normalized
from update_accessories_period_incremental import refresh_metadata
from update_cruces_prices_from_lists import parse_date_from_name


ACCESSORY_MODEL_OVERRIDE_SOURCE = {
    "Generico Cargador Inalambrico Nature": {"TIPO": "ENERGIA", "MARCA": "GENERICO"},
    "Honor Tab X8B LTE": {"TIPO": "TABLET", "MARCA": "HONOR"},
    "Generico Lamina Vidrio Zte Nubia Air 5G": {"TIPO": "PROTECCION", "MARCA": "GENERICO"},
    "Huawei Matepad 11.5 256GB": {"TIPO": "TABLET", "MARCA": "HUAWEI"},
    "Huawei Matepad SE11 128GB LTE": {"TIPO": "TABLET", "MARCA": "HUAWEI"},
    "Generico Case Iphone 17 Pro Max": {"TIPO": "PROTECCION", "MARCA": "GENERICO"},
    "Generico Lamina Vidrio Iphone 17 Pro Max": {"TIPO": "PROTECCION", "MARCA": "GENERICO"},
    "Huawei Matepad 12X 256GB": {"TIPO": "TABLET", "MARCA": "HUAWEI"},
    "Nintendo Switch Oled + Super Mario Wonder GAMING": {"TIPO": "GAMING", "MARCA": "NINTENDO"},
    "Nintendo Switch Oled + Super Mario Wonder": {"TIPO": "GAMING", "MARCA": "NINTENDO"},
    "Xiaomi Redmi Pad 2 9 Pulgadas": {"TIPO": "TABLET", "MARCA": "XIAOMI"},
    "Huawei Watch Fit 5 Pro": {"TIPO": "WEAREABLES", "MARCA": "HUAWEI"},
}
ACCESSORY_MODEL_OVERRIDES = {norm_key(model): value for model, value in ACCESSORY_MODEL_OVERRIDE_SOURCE.items()}

MANUAL_PRICE_ALIASES = {
    norm_key("Altrac Promocional Tracker Tag"): "ALTRAC TRACKER TAG PROMO WT",
    norm_key("Apple Audifonos Earpods USB C"): "APPLE EARPODS CONECTOR TYPE C",
    norm_key("Generico Cargador Inalambrico Nature"): "GENERICO CARG INAL NATURE USB DUAL 6534",
    norm_key("Generico Case Iphone 17 Pro Max"): "GENERICO CASE IPHONE 16 PRO MX",
    norm_key("Generico Lamina Vidrio Iphone 17 Pro Max"): "GENERICO LAM IPHONE 16 PRO MX",
    norm_key("Generico Lamina Vidrio Zte Nubia Air 5G"): "GYRUX LAMINA VIDRIO BLADE A51",
    norm_key("Generico Promocional Parlante"): "MOW PARLANTE R2",
    norm_key("Maxell Promocional Audifono Over Ear"): "MAXELL PROMO AUDIF OE NEGRO",
    norm_key("Nintendo Switch Oled + Super Mario Wonder GAMING"): "NINTENDO SWITCH OLED BUNDLE SMWONDER",
    norm_key("Nintendo Switch Oled + Super Mario Wonder"): "NINTENDO SWITCH OLED BUNDLE SMWONDER",
    norm_key("Xiaomi Powerbank 10.000 mAh Lite"): "XIAOMI POWRBNK 10KMAH LITE WT",
    norm_key("Xiaomi Redmi Powerbank 20.000 mAh"): "XIAOMI RDM POWERBANK 20KMAH BK",
}

ACCESSORY_PRICE_SHEET = "Accesorios_entel_TP_TPF"
PRICE_HEADER_ALIASES = {"precio actual", "precioactual"}
DESCRIPTION_HEADER_ALIASES = {"descripcion", "descripción", "descripcion producto", "descripción producto"}

TOKEN_STOPWORDS = {
    "accesorio",
    "accesorios",
    "acc",
    "generico",
    "generica",
    "color",
    "modelo",
    "gaming",
    "negro",
    "negr",
    "black",
    "blanco",
    "white",
    "azul",
    "blue",
    "verde",
    "green",
    "rojo",
    "red",
    "gris",
    "gray",
    "grey",
    "plata",
    "silver",
    "dorado",
    "gold",
    "rosado",
    "rosa",
    "pink",
    "morado",
    "purpura",
    "purple",
    "transparente",
    "clear",
}


def ascii_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", normalized(value).casefold())
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def compact_key(value: str) -> str:
    return ascii_text(value).replace(" ", "")


def model_tokens(value: str) -> list[str]:
    return [token for token in ascii_text(value).split() if token and token not in TOKEN_STOPWORDS]


def numeric_price(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalized(value).replace("S/", "").replace(",", "").strip()
    if not text or text in {"-", "NA", "N/A"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_iso_date(value: str | None) -> date | None:
    if not value or value == "Sin fecha":
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def find_header(ws) -> tuple[int, int, int]:
    for row_idx in range(1, min(ws.max_row or 0, 80) + 1):
        headers = {
            ascii_text(ws.cell(row_idx, column).value): column
            for column in range(1, (ws.max_column or 0) + 1)
            if normalized(ws.cell(row_idx, column).value)
        }
        description_col = next((headers[key] for key in DESCRIPTION_HEADER_ALIASES if key in headers), None)
        price_col = next((headers[key] for key in PRICE_HEADER_ALIASES if key in headers), None)
        if description_col and price_col:
            return row_idx, description_col, price_col
    raise RuntimeError(f"No encontré encabezados de accesorios en {ws.title}.")


def extract_price_candidates(path: Path) -> list[dict]:
    list_date = parse_date_from_name(path)
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        if ACCESSORY_PRICE_SHEET not in wb.sheetnames:
            return []
        ws = wb[ACCESSORY_PRICE_SHEET]
        header_row, description_col, price_col = find_header(ws)
        rows = []
        seen = {}
        for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
            description = normalized(ws.cell(row_idx, description_col).value)
            price = numeric_price(ws.cell(row_idx, price_col).value)
            if not description or price is None:
                continue
            key = compact_key(description)
            item = {
                "date": list_date,
                "file": path.name,
                "description": description,
                "price": price,
                "compact": key,
                "tokens": model_tokens(description),
            }
            seen[key] = item
        rows.extend(seen.values())
        return rows
    finally:
        wb.close()


def collect_price_files(paths: list[Path]) -> list[Path]:
    candidates = []
    for path in paths:
        if path.is_dir():
            candidates.extend(
                item
                for item in path.iterdir()
                if item.is_file() and item.suffix.lower() in {".xlsm", ".xlsx"}
            )
        elif path.exists():
            candidates.append(path)
    by_date: dict[date, Path] = {}
    for path in candidates:
        if path.name.startswith("._"):
            continue
        try:
            list_date = parse_date_from_name(path)
        except ValueError:
            continue
        current = by_date.get(list_date)
        if current is None or path.stat().st_mtime >= current.stat().st_mtime:
            by_date[list_date] = path
    return [by_date[key] for key in sorted(by_date)]


def price_files_for_rows(price_files: list[Path], rows: list[dict]) -> list[Path]:
    row_dates = [parse_iso_date(row.get("FECHA")) for row in rows]
    row_dates = [value for value in row_dates if value is not None]
    if not row_dates or not price_files:
        return price_files
    first_sale_date = min(row_dates)
    dated = [(parse_date_from_name(path), path) for path in price_files]
    before = [item for item in dated if item[0] <= first_sale_date]
    selected = [path for value, path in dated if value >= first_sale_date]
    if before:
        selected.insert(0, max(before, key=lambda item: item[0])[1])
    return list(dict.fromkeys(selected))


def build_price_catalog(price_files: list[Path]) -> dict:
    candidates_by_date = defaultdict(list)
    sources = []
    for path in price_files:
        candidates = extract_price_candidates(path)
        if not candidates:
            continue
        list_date = candidates[0]["date"]
        candidates_by_date[list_date].extend(candidates)
        sources.append({"date": list_date.isoformat(), "file": str(path), "rows": len(candidates)})
    dates = sorted(candidates_by_date)
    return {"dates": dates, "by_date": dict(candidates_by_date), "sources": sources}


def list_date_for_row(catalog: dict, row_date: date | None) -> date | None:
    dates = catalog["dates"]
    if not dates:
        return None
    if row_date is None:
        return dates[-1]
    if row_date < dates[0]:
        return dates[-1]
    available = [value for value in dates if value <= row_date]
    return available[-1] if available else dates[-1]


def match_score(query: str, candidate: dict) -> float:
    query_compact = compact_key(query)
    if query_compact and query_compact == candidate["compact"]:
        return 1.0
    query_tokens = model_tokens(query)
    candidate_tokens = candidate["tokens"]
    if not query_tokens or not candidate_tokens:
        return 0.0
    query_set = set(query_tokens)
    candidate_set = set(candidate_tokens)
    overlap = len(query_set & candidate_set) / len(query_set)
    reverse_overlap = len(query_set & candidate_set) / len(candidate_set)
    query_text = " ".join(query_tokens)
    candidate_text = " ".join(candidate_tokens)
    ratio = SequenceMatcher(None, query_text, candidate_text).ratio()
    if query_compact and query_compact in candidate["compact"]:
        return max(0.96, ratio)
    if candidate["compact"] and candidate["compact"] in query_compact and len(candidate_tokens) >= 3:
        return max(0.9, ratio)
    if query_set <= candidate_set:
        return max(0.95, ratio)
    if candidate_set <= query_set and len(candidate_set) >= min(3, len(query_set)):
        return max(0.88, ratio)
    return max(ratio, overlap * 0.78 + reverse_overlap * 0.22)


def best_price_match(model: str, candidates: list[dict]) -> dict | None:
    alias = MANUAL_PRICE_ALIASES.get(norm_key(model))
    if alias:
        alias_compact = compact_key(alias)
        for candidate in candidates:
            if alias_compact == candidate["compact"] or alias_compact in candidate["compact"] or candidate["compact"] in alias_compact:
                return {**candidate, "score": 1.0}

    best = None
    best_score = 0.0
    for candidate in candidates:
        score = match_score(model, candidate)
        if score > best_score:
            best = candidate
            best_score = score
    if best is None:
        return None
    query_tokens = set(model_tokens(model))
    candidate_tokens = set(best["tokens"])
    overlap = len(query_tokens & candidate_tokens) / len(query_tokens) if query_tokens else 0.0
    if best_score >= 0.72 and overlap >= 0.5:
        return {**best, "score": best_score}
    return None


def apply_model_overrides(row: dict, changes: Counter) -> None:
    override = ACCESSORY_MODEL_OVERRIDES.get(norm_key(row.get("MARCAMODELO")))
    if not override:
        return
    for field, value in override.items():
        if row.get(field) != value:
            row[field] = value
            changes[f"{field.lower()}_rows"] += 1


def apply_category_rules(row: dict, changes: Counter) -> None:
    if normalized(row.get("TIPO")).casefold() == "tablet" and row.get("SUBTIPO") != "TABLET":
        row["SUBTIPO"] = "TABLET"
        row["SEGMENTO_HONOR"] = "TABLET"
        changes["tablet_subtype_rows"] += 1


def apply_regional_overrides(row: dict, changes: Counter) -> None:
    regional = {norm_key(key): value for key, value in REGIONAL_OVERRIDES.items()}.get(norm_key(row.get("PUNTODEVENTA")))
    if regional and row.get("REGIONAL_HONOR") != regional:
        row["REGIONAL_HONOR"] = regional
        changes["regional_rows"] += 1


def apply_channel_group(row: dict, changes: Counter) -> None:
    channel = row.get("CANAL_AGRUPADO")
    grouped = group_channel(channel)
    if channel != grouped:
        row["CANAL_AGRUPADO"] = grouped
        changes["channel_rows"] += 1


def apply_prices(rows: list[dict], catalog: dict) -> Counter:
    changes = Counter()
    cache = {}
    for row in rows:
        model = row.get("MARCAMODELO")
        row_date = parse_iso_date(row.get("FECHA"))
        effective_date = list_date_for_row(catalog, row_date)
        cache_key = (norm_key(model), effective_date)
        if cache_key not in cache:
            candidates = catalog["by_date"].get(effective_date, []) if effective_date else []
            cache[cache_key] = best_price_match(model, candidates) if model else None
        match = cache[cache_key]
        if match:
            row["PRECIO"] = match["price"]
            row["PRECIO_MATCH"] = match["description"]
            row["PRECIO_MATCH_SCORE"] = round(match["score"], 4)
            row["PRECIO_LISTA_FECHA"] = match["date"].isoformat()
            changes["price_rows"] += 1
        else:
            row["PRECIO"] = None
            row["PRECIO_MATCH"] = None
            row["PRECIO_MATCH_SCORE"] = None
            row["PRECIO_LISTA_FECHA"] = effective_date.isoformat() if effective_date else None
            changes["missing_price_rows"] += 1
    return changes


def recompute_price_summary(data: dict) -> None:
    missing_price = Counter()
    matched_models = {}
    for row in data["dashboard_data"]:
        sales = float(row.get("VENTAS", 0) or 0)
        if row.get("PRECIO") is None:
            missing_price[row["MARCAMODELO"]] += sales
        elif row["MARCAMODELO"] not in matched_models:
            matched_models[row["MARCAMODELO"]] = {
                "PRECIO": row.get("PRECIO"),
                "PRECIO_MATCH": row.get("PRECIO_MATCH"),
                "PRECIO_MATCH_SCORE": row.get("PRECIO_MATCH_SCORE"),
                "PRECIO_LISTA_FECHA": row.get("PRECIO_LISTA_FECHA"),
            }
    summary = data.setdefault("cross_summary", {})
    missing_sales = summary.setdefault("missing_sales", {})
    if missing_price:
        missing_sales["PRECIO"] = round(sum(missing_price.values()), 6)
    else:
        missing_sales.pop("PRECIO", None)
    summary["missing_price_models"] = [
        {"MARCAMODELO": model, "VENTAS": round(sales, 6)}
        for model, sales in missing_price.most_common(80)
    ]
    summary["matched_price_models"] = len(matched_models)


def enrich(input_path: Path, output_path: Path, price_paths: list[Path], cruces: str | None = None) -> dict:
    data = json.loads(input_path.read_text(encoding="utf-8"))
    changes = Counter()
    for row in data.get("dashboard_data", []):
        apply_channel_group(row, changes)
        apply_model_overrides(row, changes)
        apply_category_rules(row, changes)
        apply_regional_overrides(row, changes)
    for row in data.get("stock_data", []):
        apply_channel_group(row, changes)
        apply_model_overrides(row, changes)
        apply_category_rules(row, changes)
        apply_regional_overrides(row, changes)

    price_files = price_files_for_rows(
        collect_price_files(price_paths),
        data.get("dashboard_data", []),
    )
    catalog = build_price_catalog(price_files)
    changes.update(apply_prices(data.get("dashboard_data", []), catalog))

    if cruces:
        data["cruces_file"] = cruces
    data["source_file"] = "Bases Oracle 2024-2026 - Familia Accesorios"
    data["generated_at"] = datetime.now().isoformat(timespec="seconds")
    data["accessory_price_sources"] = catalog["sources"]
    refresh_metadata(data)
    recompute_price_summary(data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "output": str(output_path),
        "changes": dict(changes),
        "price_sources": catalog["sources"],
        "missing_price_models": data.get("cross_summary", {}).get("missing_price_models", [])[:20],
        "missing_crosses": data.get("cross_summary", {}).get("missing_sales", {}),
        "total_sales": data.get("total_sales"),
        "periods": data.get("lists", {}).get("periods", []),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Enriquece el reporte de accesorios con correcciones, regionales y precios.")
    parser.add_argument("input")
    parser.add_argument("output")
    parser.add_argument("--price-path", action="append", type=Path, required=True)
    parser.add_argument("--cruces")
    args = parser.parse_args()
    result = enrich(Path(args.input), Path(args.output), args.price_path, args.cruces)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
