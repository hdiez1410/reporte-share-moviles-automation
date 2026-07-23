from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def normalized(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def norm_key(value) -> str:
    text = unicodedata.normalize("NFKD", normalized(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.casefold().split())


def read_overrides(path: Path) -> dict[str, dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not {"MARCAMODELO", "Segmento"}.issubset(reader.fieldnames or []):
            raise RuntimeError("El CSV debe tener columnas MARCAMODELO y Segmento.")
        overrides = {}
        for row in reader:
            model = normalized(row.get("MARCAMODELO"))
            segment = normalized(row.get("Segmento"))
            if not model or not segment:
                continue
            overrides[norm_key(model)] = {"MARCAMODELO": model, "Segmento": segment}
        return overrides


def apply_overrides(cruces: Path, overrides_csv: Path, output: Path, audit: Path) -> dict:
    overrides = read_overrides(overrides_csv)
    wb = load_workbook(cruces)
    if "Segmento" not in wb.sheetnames:
        raise RuntimeError("CRUCES no tiene hoja Segmento.")
    ws = wb["Segmento"]

    rows_by_key = {}
    for row_idx in range(2, ws.max_row + 1):
        model = normalized(ws.cell(row_idx, 1).value)
        if model:
            rows_by_key[norm_key(model)] = row_idx

    updated = 0
    unchanged = 0
    appended = 0
    for key, item in sorted(overrides.items(), key=lambda kv: kv[1]["MARCAMODELO"].casefold()):
        row_idx = rows_by_key.get(key)
        if row_idx:
            current = normalized(ws.cell(row_idx, 2).value)
            if current != item["Segmento"]:
                ws.cell(row_idx, 2).value = item["Segmento"]
                updated += 1
            else:
                unchanged += 1
        else:
            ws.append([item["MARCAMODELO"], item["Segmento"]])
            appended += 1

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    summary = {
        "cruces_source": str(cruces),
        "cruces_output": str(output),
        "overrides_csv": str(overrides_csv),
        "overrides": len(overrides),
        "updated": updated,
        "unchanged": unchanged,
        "appended": appended,
    }
    audit.parent.mkdir(parents=True, exist_ok=True)
    audit.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Aplica overrides de Segmento Honor sobre CRUCES.")
    parser.add_argument("--cruces", required=True)
    parser.add_argument("--overrides", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--audit", required=True)
    args = parser.parse_args()
    summary = apply_overrides(Path(args.cruces), Path(args.overrides), Path(args.output), Path(args.audit))
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
