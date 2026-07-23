from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


REGIONAL_OVERRIDES = {
    "Sin punto de venta": "OTROS",
    "WONG_ASIA2": "CLEIBER CORTEZ",
    "TEXPRESS_AQPVIDAURRAZAGA": "TEPHY DIAZ",
    "TEXPRESS_PUNAYAVIRI": "TEPHY DIAZ",
    "TEX_SATELITE_CUSTICATICA": "TEPHY DIAZ",
    "HUB_AREQUIPA": "TEPHY DIAZ",
    "TE PIURA LORETO": "LILIANA BONILLA",
    "TE CHINCHA CARRION": "WENDY CAWEN",
    "TE PAITA": "LILIANA BONILLA",
    "TE AQP SIGLO XX": "TEPHY DIAZ",
    "TE JUN TARMA": "WENDY CAWEN",
    "TE PRESID HUANC": "WENDY CAWEN",
    "TE ANDAHUAYLAS MARTINELLI": "TEPHY DIAZ",
    "TE JERUSALEN 2": "TEPHY DIAZ",
    "TE NAJAR AQP": "TEPHY DIAZ",
    "TE PIURA CALLAO": "LILIANA BONILLA",
    "TE LIM PALERMO VICTORIA": "CLEIBER CORTEZ",
    "TE CIX BOLOGNESI": "LILIANA BONILLA",
    "TE LIM BELLAVISTA CALLAO": "CLEIBER CORTEZ",
    "TE LIM CULTURA AVIACION": "CLEIBER CORTEZ",
    "TE SMP FILADELFIA": "CLEIBER CORTEZ",
    "BACK ANULACION": "OTROS",
}

SEGMENT_OVERRIDES = {
    "Apple Iphone 14 Pro Max 512GB": "Apple",
    "Huawei Nova Y60": "Sin Segmento",
    "IS 655.2": "Sin Segmento",
    "Oppo Reno 13F 512GB 5G": "Honor 600 E",
    "Xiaomi Redmi A7 Pro 128GB 4G": "Honor X5C Plus",
}


def main() -> int:
    parser = argparse.ArgumentParser(description="Agrega PDV faltantes a CRUCES/REGIONAL.")
    parser.add_argument("input")
    parser.add_argument("output")
    args = parser.parse_args()

    wb = load_workbook(args.input)
    regional_ws = wb["REGIONAL"]
    regional_existing = {}
    for row in regional_ws.iter_rows(min_row=2, max_col=2):
        if row[0].value:
            regional_existing[str(row[0].value).strip().casefold()] = row
    added = []
    updated = []
    for pdv, regional in REGIONAL_OVERRIDES.items():
        existing_row = regional_existing.get(pdv.casefold())
        if existing_row:
            if str(existing_row[1].value).strip() != regional:
                existing_row[1].value = regional
                updated.append({"PDV": pdv, "Regional Honor": regional})
            continue
        regional_ws.append([pdv, regional])
        added.append({"PDV": pdv, "Regional Honor": regional})

    segment_ws = wb["Segmento"]
    segment_existing = {}
    for row in segment_ws.iter_rows(min_row=2, max_col=2):
        if row[0].value:
            segment_existing[str(row[0].value).strip().casefold()] = row
    added_segments = []
    updated_segments = []
    for model, segment in SEGMENT_OVERRIDES.items():
        existing_row = segment_existing.get(model.casefold())
        if existing_row:
            if str(existing_row[1].value).strip() != segment:
                existing_row[1].value = segment
                updated_segments.append({"MARCAMODELO": model, "Segmento": segment})
            continue
        segment_ws.append([model, segment])
        added_segments.append({"MARCAMODELO": model, "Segmento": segment})

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(
        {
            "added_regional": added,
            "updated_regional": updated,
            "added_segment": added_segments,
            "updated_segment": updated_segments,
            "output": str(out),
        }
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
